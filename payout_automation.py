"""
payout_automation.py
FinOps Seller Payout Intelligence System — Reconciliation Pipeline
Amazon Seller Services · Process Associate L3

Pipeline Steps:
  [1/6] Monthly payout reconciliation
  [2/6] Root cause classification engine
  [3/6] SLA breach tracker
  [4/6] Seller risk scoring & prediction
  [5/6] Category performance analysis
  [6/6] Stakeholder digest generation (Excel + reports)
"""

import pandas as pd
import numpy as np
import sqlite3
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

DB_PATH = "data/seller_payouts.db"
REPORT_DIR = "reports"
os.makedirs(REPORT_DIR, exist_ok=True)

# ─── COLORS ──────────────────────────────────────────────────────────────────
C_AMAZON_ORANGE = "FF9900"
C_AMAZON_DARK   = "131921"
C_GREEN         = "00A651"
C_RED           = "D0021B"
C_AMBER         = "F5A623"
C_BLUE          = "0066C0"
C_LIGHT_GRAY    = "F3F3F3"
C_MID_GRAY      = "CCCCCC"
C_WHITE         = "FFFFFF"
C_DARK_GREEN    = "005C35"
C_DARK_RED      = "7B0000"


# ─── HELPERS ─────────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color=C_MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def pct(val, total):
    return round(val / total * 100, 1) if total else 0

def inr(val):
    return f"₹{val:,.0f}"


# ─── LOAD DATA ────────────────────────────────────────────────────────────────
def load_data():
    conn = sqlite3.connect(DB_PATH)
    txn = pd.read_sql("SELECT * FROM transactions", conn)
    sellers = pd.read_sql("SELECT * FROM sellers", conn)
    conn.close()
    return txn, sellers


# ─── [1/6] MONTHLY RECONCILIATION ────────────────────────────────────────────
def monthly_reconciliation(txn, sellers):
    print("\n[1/6] Running monthly payout reconciliation...")

    monthly = txn.groupby("month").agg(
        active_sellers=("seller_id", "nunique"),
        total_transactions=("transaction_id", "count"),
        total_gmv=("gmv", "sum"),
        total_commission=("commission", "sum"),
        total_gst=("gst_on_commission", "sum"),
        total_returns=("return_amount", "sum"),
        total_expected_payout=("expected_payout", "sum"),
        total_actual_payout=("actual_payout", "sum"),
        discrepancy_count=("status", lambda x: (x == "Discrepancy").sum()),
        discrepancy_amount=("discrepancy_amount", lambda x: x.abs().sum()),
        sla_breaches=("sla_breach", "sum"),
    ).reset_index()

    monthly["reconciliation_gap"] = (
        monthly["total_actual_payout"] - monthly["total_expected_payout"]
    ).round(2)
    monthly["discrepancy_rate"] = (
        monthly["discrepancy_count"] / monthly["total_transactions"] * 100
    ).round(2)

    # MoM growth
    monthly = monthly.sort_values("month")
    monthly["gmv_mom_growth"] = monthly["total_gmv"].pct_change() * 100
    monthly["payout_mom_growth"] = monthly["total_expected_payout"].pct_change() * 100

    print(f"   ✅ {len(monthly)} months reconciled")
    for _, row in monthly.iterrows():
        status = "⚠️ " if abs(row["reconciliation_gap"]) > 5000 else "✅"
        print(f"   {status} {row['month']}: GMV={inr(row['total_gmv'])} | "
              f"Payout={inr(row['total_actual_payout'])} | "
              f"Discrepancies={row['discrepancy_count']} ({row['discrepancy_rate']}%)")

    return monthly


# ─── [2/6] ROOT CAUSE ENGINE ─────────────────────────────────────────────────
def root_cause_analysis(txn):
    print("\n[2/6] Running root cause classification engine...")

    disc = txn[txn["status"] == "Discrepancy"].copy()

    rc_summary = disc.groupby("root_cause").agg(
        count=("transaction_id", "count"),
        total_amount=("discrepancy_amount", lambda x: x.abs().sum()),
        avg_amount=("discrepancy_amount", lambda x: x.abs().mean()),
        overpayments=("discrepancy_amount", lambda x: (x > 0).sum()),
        underpayments=("discrepancy_amount", lambda x: (x < 0).sum()),
    ).reset_index()

    rc_summary["pct_of_total"] = (
        rc_summary["count"] / rc_summary["count"].sum() * 100
    ).round(1)
    rc_summary["amount_pct"] = (
        rc_summary["total_amount"] / rc_summary["total_amount"].sum() * 100
    ).round(1)
    rc_summary = rc_summary.sort_values("total_amount", ascending=False)

    # Severity classification
    def classify_severity(row):
        if row["total_amount"] > 30000 or row["count"] > 80:
            return "CRITICAL"
        elif row["total_amount"] > 15000 or row["count"] > 40:
            return "HIGH"
        elif row["total_amount"] > 5000 or row["count"] > 15:
            return "MEDIUM"
        return "LOW"

    rc_summary["severity"] = rc_summary.apply(classify_severity, axis=1)

    print(f"   ✅ {len(disc)} discrepancies classified into {len(rc_summary)} root causes")
    for _, row in rc_summary.iterrows():
        icon = {"CRITICAL": "🔴", "HIGH": "🟠", "MEDIUM": "🟡", "LOW": "🟢"}[row["severity"]]
        print(f"   {icon} {row['root_cause']}: {row['count']} cases | "
              f"{inr(row['total_amount'])} | {row['pct_of_total']}%")

    return rc_summary, disc


# ─── [3/6] SLA BREACH TRACKER ────────────────────────────────────────────────
def sla_tracker(txn, sellers):
    print("\n[3/6] Running SLA breach tracker...")

    breaches = txn[txn["sla_breach"] == True].copy()
    breaches = breaches.merge(
        sellers[["seller_id", "seller_name", "tier"]], on="seller_id", how="left"
    )

    # Aging buckets
    def age_bucket(days):
        if days <= 2:   return "0-2 days (OK)"
        elif days <= 7: return "3-7 days (Warning)"
        elif days <= 14: return "8-14 days (Overdue)"
        else:           return "15+ days (Critical)"

    breaches["aging_bucket"] = breaches["days_outstanding"].apply(age_bucket)

    aging_summary = breaches.groupby("aging_bucket").agg(
        count=("transaction_id", "count"),
        amount_at_risk=("discrepancy_amount", lambda x: x.abs().sum()),
    ).reset_index()

    critical_breaches = breaches[breaches["days_outstanding"] > 14].sort_values(
        "discrepancy_amount", key=abs, ascending=False
    ).head(20)

    print(f"   ✅ {len(breaches)} SLA breaches tracked")
    for _, row in aging_summary.iterrows():
        print(f"   📅 {row['aging_bucket']}: {row['count']} cases | {inr(row['amount_at_risk'])} at risk")

    return aging_summary, critical_breaches, breaches


# ─── [4/6] SELLER RISK SCORING ───────────────────────────────────────────────
def seller_risk_scoring(txn, sellers):
    print("\n[4/6] Computing seller risk scores & next-month predictions...")

    # Aggregate per seller
    seller_metrics = txn.groupby("seller_id").agg(
        total_gmv=("gmv", "sum"),
        total_transactions=("transaction_id", "count"),
        discrepancy_count=("status", lambda x: (x == "Discrepancy").sum()),
        discrepancy_amount=("discrepancy_amount", lambda x: x.abs().sum()),
        sla_breaches=("sla_breach", "sum"),
        avg_days_outstanding=("days_outstanding", "mean"),
        unique_root_causes=("root_cause", "nunique"),
    ).reset_index()

    seller_metrics = seller_metrics.merge(
        sellers[["seller_id", "seller_name", "tier", "commission_rate",
                 "return_rate", "risk_score", "state", "category"]], on="seller_id"
    )

    seller_metrics["disc_rate"] = (
        seller_metrics["discrepancy_count"] / seller_metrics["total_transactions"] * 100
    ).round(2)

    # Computed risk score (0-100, higher = riskier)
    def compute_risk(row):
        score = 0
        score += min(40, row["disc_rate"] * 4)          # discrepancy rate (max 40)
        score += min(20, row["sla_breaches"] * 1.5)      # SLA breaches (max 20)
        score += min(15, row["unique_root_causes"] * 3)  # variety of issues (max 15)
        score += min(15, row["return_rate"] * 80)        # return rate (max 15)
        score += min(10, row["avg_days_outstanding"] / 5) # days outstanding (max 10)
        return round(min(100, score), 1)

    seller_metrics["computed_risk_score"] = seller_metrics.apply(compute_risk, axis=1)

    def risk_tier(score):
        if score >= 70: return "CRITICAL"
        elif score >= 50: return "HIGH"
        elif score >= 30: return "MEDIUM"
        return "LOW"

    seller_metrics["risk_tier"] = seller_metrics["computed_risk_score"].apply(risk_tier)

    # Next month prediction flag (logistic-style rule)
    seller_metrics["predicted_anomaly_next_month"] = (
        (seller_metrics["disc_rate"] > 10) |
        (seller_metrics["sla_breaches"] > 5) |
        (seller_metrics["computed_risk_score"] > 60)
    )

    high_risk = seller_metrics[seller_metrics["risk_tier"].isin(["CRITICAL", "HIGH"])].sort_values(
        "computed_risk_score", ascending=False
    )

    predicted_flag_count = seller_metrics["predicted_anomaly_next_month"].sum()
    print(f"   ✅ Risk scores computed for {len(seller_metrics)} sellers")
    print(f"   🔴 Critical risk: {len(seller_metrics[seller_metrics['risk_tier']=='CRITICAL'])} sellers")
    print(f"   🟠 High risk:     {len(seller_metrics[seller_metrics['risk_tier']=='HIGH'])} sellers")
    print(f"   📊 Predicted anomalies next month: {predicted_flag_count} sellers")

    return seller_metrics, high_risk


# ─── [5/6] CATEGORY PERFORMANCE ──────────────────────────────────────────────
def category_analysis(txn):
    print("\n[5/6] Analyzing category performance...")

    cat = txn.groupby("category").agg(
        total_gmv=("gmv", "sum"),
        total_transactions=("transaction_id", "count"),
        total_payout=("actual_payout", "sum"),
        discrepancies=("status", lambda x: (x == "Discrepancy").sum()),
        disc_amount=("discrepancy_amount", lambda x: x.abs().sum()),
        avg_return_rate=("return_amount", lambda x: (x / txn.loc[x.index, "gmv"]).mean()),
    ).reset_index()

    cat["disc_rate"] = (cat["discrepancies"] / cat["total_transactions"] * 100).round(2)
    cat["avg_transaction"] = (cat["total_gmv"] / cat["total_transactions"]).round(2)
    cat = cat.sort_values("total_gmv", ascending=False)

    print(f"   ✅ {len(cat)} categories analyzed")
    return cat


# ─── [6/6] EXCEL REPORT ──────────────────────────────────────────────────────
def build_excel_report(monthly, rc_summary, aging_summary, critical_breaches,
                        seller_metrics, high_risk, category_data, txn):
    print("\n[6/6] Building stakeholder Excel report...")

    wb = Workbook()

    # ── Sheet styles ──────────────────────────────────────────────────────────
    HEADER_FONT = Font(name="Calibri", bold=True, color=C_WHITE, size=11)
    TITLE_FONT  = Font(name="Calibri", bold=True, color=C_AMAZON_DARK, size=14)
    MONO_FONT   = Font(name="Courier New", size=10)
    NORMAL_FONT = Font(name="Calibri", size=10)
    BOLD_FONT   = Font(name="Calibri", bold=True, size=10)

    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left", vertical="center")
    RIGHT  = Alignment(horizontal="right", vertical="center")
    WRAP   = Alignment(wrap_text=True, vertical="center")

    def style_header_row(ws, row_num, col_count, fill_color, font=None):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = header_fill(fill_color)
            cell.font = font or HEADER_FONT
            cell.alignment = CENTER
            cell.border = thin_border()

    def style_data_row(ws, row_num, col_count, fill_color=None, font=None, align=None):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            if fill_color:
                cell.fill = header_fill(fill_color)
            cell.font = font or NORMAL_FONT
            cell.alignment = align or LEFT
            cell.border = thin_border()

    def add_title_block(ws, title, subtitle, row=1):
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = TITLE_FONT
        ws[f"A{row}"].fill = header_fill("FFF3CD")
        ws[f"A{row}"].alignment = CENTER

        ws.merge_cells(f"A{row+1}:H{row+1}")
        ws[f"A{row+1}"] = subtitle
        ws[f"A{row+1}"].font = Font(name="Calibri", color="555555", italic=True, size=10)
        ws[f"A{row+1}"].alignment = CENTER
        return row + 3

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1: EXECUTIVE SUMMARY
    # ─────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = C_AMAZON_ORANGE

    # Title
    ws1.merge_cells("A1:H2")
    ws1["A1"] = "🏢  FINOPS SELLER PAYOUT INTELLIGENCE SYSTEM"
    ws1["A1"].font = Font(name="Calibri", bold=True, color=C_WHITE, size=16)
    ws1["A1"].fill = header_fill(C_AMAZON_DARK)
    ws1["A1"].alignment = CENTER

    ws1.merge_cells("A3:H3")
    ws1["A3"] = f"Amazon Seller Services · Q4 2024 (Oct–Dec) · Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws1["A3"].font = Font(name="Calibri", italic=True, color=C_WHITE, size=10)
    ws1["A3"].fill = header_fill("1A252F")
    ws1["A3"].alignment = CENTER

    ws1.row_dimensions[1].height = 28
    ws1.row_dimensions[3].height = 18

    # KPI Cards (row 5 onwards)
    ws1["A5"] = "KEY PERFORMANCE INDICATORS"
    ws1["A5"].font = Font(bold=True, color=C_AMAZON_DARK, size=12)
    ws1.merge_cells("A5:H5")
    ws1["A5"].alignment = CENTER

    kpis = [
        ("Total GMV",           inr(txn["gmv"].sum()),                        C_AMAZON_DARK),
        ("Net Payout",          inr(txn["actual_payout"].sum()),               "005C35"),
        ("Discrepancies",       f"{(txn['status']=='Discrepancy').sum()}",      C_RED),
        ("Amount at Risk",      inr(txn[txn['status']=='Discrepancy']['discrepancy_amount'].abs().sum()), C_RED),
        ("SLA Breaches",        f"{txn['sla_breach'].sum()}",                  C_AMBER),
        ("Active Sellers",      f"{txn['seller_id'].nunique()}",               C_BLUE),
        ("Avg Disc. Rate",      f"{(txn['status']=='Discrepancy').mean()*100:.1f}%", C_AMBER),
        ("Months Covered",      "3 (Q4 2024)",                                 C_AMAZON_DARK),
    ]

    ws1.row_dimensions[6].height = 14
    ws1.row_dimensions[7].height = 30
    ws1.row_dimensions[8].height = 30
    ws1.row_dimensions[9].height = 14

    for i, (label, value, color) in enumerate(kpis):
        col = i + 1
        ws1.column_dimensions[get_column_letter(col)].width = 18
        # Label
        lc = ws1.cell(row=7, column=col, value=label)
        lc.font = Font(name="Calibri", bold=True, color=C_WHITE, size=9)
        lc.fill = header_fill(color)
        lc.alignment = CENTER
        lc.border = thin_border()
        # Value
        vc = ws1.cell(row=8, column=col, value=value)
        vc.font = Font(name="Calibri", bold=True, size=12)
        vc.fill = header_fill("F8F8F8")
        vc.alignment = CENTER
        vc.border = thin_border()

    # Monthly Table
    start_row = 11
    ws1[f"A{start_row}"] = "MONTHLY RECONCILIATION SUMMARY"
    ws1[f"A{start_row}"].font = Font(bold=True, size=11, color=C_WHITE)
    ws1[f"A{start_row}"].fill = header_fill(C_AMAZON_DARK)
    ws1.merge_cells(f"A{start_row}:H{start_row}")
    ws1[f"A{start_row}"].alignment = CENTER
    ws1.row_dimensions[start_row].height = 20

    hdr = start_row + 1
    monthly_headers = ["Month", "Active Sellers", "Transactions", "Total GMV",
                        "Net Payout", "Discrepancies", "Amount at Risk", "SLA Breaches"]
    for ci, h in enumerate(monthly_headers, 1):
        c = ws1.cell(row=hdr, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill(C_AMAZON_ORANGE)
        c.alignment = CENTER
        c.border = thin_border()

    for ri, (_, row) in enumerate(monthly.iterrows(), hdr + 1):
        vals = [row["month"], row["active_sellers"], row["total_transactions"],
                row["total_gmv"], row["total_actual_payout"], row["discrepancy_count"],
                row["discrepancy_amount"], row["sla_breaches"]]
        fill = "FFFFFF" if ri % 2 == 0 else C_LIGHT_GRAY
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci)
            if isinstance(val, float):
                c.value = round(val, 2)
                c.number_format = '₹#,##0.00' if ci in [4,5,7] else '0.00'
            else:
                c.value = val
            c.font = NORMAL_FONT
            c.fill = header_fill(fill)
            c.alignment = CENTER if ci > 1 else LEFT
            c.border = thin_border()
        ws1.row_dimensions[ri].height = 18

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2: ROOT CAUSE ANALYSIS
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Root Cause Analysis")
    ws2.sheet_properties.tabColor = C_RED

    ws2.merge_cells("A1:G2")
    ws2["A1"] = "🔍  ROOT CAUSE INTELLIGENCE ENGINE"
    ws2["A1"].font = Font(bold=True, color=C_WHITE, size=14)
    ws2["A1"].fill = header_fill(C_RED)
    ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 28

    ws2["A3"] = "Automated classification of all payout discrepancies by root cause type"
    ws2.merge_cells("A3:G3")
    ws2["A3"].font = Font(italic=True, color="555555", size=10)
    ws2["A3"].alignment = CENTER

    rc_headers = ["Root Cause Code", "Cases", "% of Total", "Amount at Risk (₹)",
                  "Avg Amount (₹)", "Overpayments", "Underpayments", "Severity"]
    hdr_row = 5
    for ci, h in enumerate(rc_headers, 1):
        c = ws2.cell(row=hdr_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill(C_RED)
        c.alignment = CENTER
        c.border = thin_border()
        ws2.column_dimensions[get_column_letter(ci)].width = 22

    severity_colors = {"CRITICAL": "D0021B", "HIGH": "F5A623", "MEDIUM": "F8E71C", "LOW": "00A651"}
    severity_text =   {"CRITICAL": C_WHITE,  "HIGH": C_WHITE,  "MEDIUM": C_AMAZON_DARK, "LOW": C_WHITE}

    for ri, (_, row) in enumerate(rc_summary.iterrows(), hdr_row + 1):
        vals = [row["root_cause"], row["count"], f"{row['pct_of_total']}%",
                round(row["total_amount"], 2), round(row["avg_amount"], 2),
                row["overpayments"], row["underpayments"], row["severity"]]
        sev = row["severity"]
        row_fill = "FFF8F8" if ri % 2 == 0 else "FFFFFF"

        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            if ci == 8:  # Severity badge
                c.fill = header_fill(severity_colors[sev])
                c.font = Font(bold=True, color=severity_text[sev], size=10)
            else:
                c.fill = header_fill(row_fill)
                c.font = Font(name="Calibri", bold=(ci == 1), size=10,
                              color=C_AMAZON_DARK if ci == 1 else "333333")
            c.alignment = CENTER
            c.border = thin_border()
            if ci in [4, 5]:
                c.number_format = '₹#,##0.00'
        ws2.row_dimensions[ri].height = 18

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 3: SLA BREACH TRACKER
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("SLA Breach Tracker")
    ws3.sheet_properties.tabColor = C_AMBER

    ws3.merge_cells("A1:H2")
    ws3["A1"] = "⏰  SLA BREACH TRACKER — CRITICAL CASES"
    ws3["A1"].font = Font(bold=True, color=C_WHITE, size=14)
    ws3["A1"].fill = header_fill("B7410E")
    ws3["A1"].alignment = CENTER
    ws3.row_dimensions[1].height = 28

    # Aging buckets summary
    ws3["A3"] = "AGING SUMMARY"
    ws3.merge_cells("A3:D3")
    ws3["A3"].font = Font(bold=True, size=11)
    ws3["A3"].fill = header_fill(C_LIGHT_GRAY)
    ws3["A3"].alignment = CENTER

    bucket_colors = {
        "0-2 days (OK)":          "D4EDDA",
        "3-7 days (Warning)":     "FFF3CD",
        "8-14 days (Overdue)":    "FFDDC1",
        "15+ days (Critical)":    "F8D7DA",
    }

    for ci, h in enumerate(["Aging Bucket", "Cases", "Amount at Risk", "Status"], 1):
        c = ws3.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill("B7410E")
        c.alignment = CENTER
        c.border = thin_border()
        ws3.column_dimensions[get_column_letter(ci)].width = 28

    for ri, (_, row) in enumerate(aging_summary.iterrows(), 5):
        bucket = row["aging_bucket"]
        bg = bucket_colors.get(bucket, "FFFFFF")
        for ci, val in enumerate([bucket, row["count"], round(row["amount_at_risk"], 2), "⚠️ Action Required" if "Critical" in bucket or "Overdue" in bucket else "✅ Monitor"], 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.fill = header_fill(bg)
            c.font = Font(name="Calibri", bold=(ci==1), size=10)
            c.alignment = CENTER
            c.border = thin_border()
            if ci == 3:
                c.number_format = '₹#,##0.00'
        ws3.row_dimensions[ri].height = 20

    # Top critical breaches
    br_start = len(aging_summary) + 7
    ws3[f"A{br_start}"] = "TOP 20 CRITICAL SLA BREACHES (15+ Days Outstanding)"
    ws3.merge_cells(f"A{br_start}:H{br_start}")
    ws3[f"A{br_start}"].font = Font(bold=True, color=C_WHITE, size=11)
    ws3[f"A{br_start}"].fill = header_fill(C_RED)
    ws3[f"A{br_start}"].alignment = CENTER

    breach_headers = ["Transaction ID", "Seller ID", "Month", "Category",
                      "Expected Payout", "Actual Payout", "Discrepancy", "Days Outstanding"]
    for ci, h in enumerate(breach_headers, 1):
        c = ws3.cell(row=br_start + 1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill(C_AMAZON_ORANGE)
        c.alignment = CENTER
        c.border = thin_border()
        ws3.column_dimensions[get_column_letter(ci)].width = 20

    for ri, (_, row) in enumerate(critical_breaches.iterrows(), br_start + 2):
        vals = [row["transaction_id"], row["seller_id"], row["month"], row["category"],
                round(row["expected_payout"], 2), round(row["actual_payout"], 2),
                round(row["discrepancy_amount"], 2), row["days_outstanding"]]
        row_fill = "FFF8F8" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.fill = header_fill(row_fill)
            c.font = NORMAL_FONT
            c.alignment = CENTER
            c.border = thin_border()
            if ci in [5, 6, 7]:
                c.number_format = '₹#,##0.00'
            if ci == 7 and isinstance(val, float):
                if val > 0:
                    c.font = Font(color="005C35", bold=True, size=10)
                else:
                    c.font = Font(color=C_RED, bold=True, size=10)
        ws3.row_dimensions[ri].height = 18

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 4: SELLER RISK SCORES
    # ─────────────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Seller Risk Scores")
    ws4.sheet_properties.tabColor = C_BLUE

    ws4.merge_cells("A1:J2")
    ws4["A1"] = "🎯  SELLER RISK INTELLIGENCE — SCORES & PREDICTIONS"
    ws4["A1"].font = Font(bold=True, color=C_WHITE, size=14)
    ws4["A1"].fill = header_fill(C_BLUE)
    ws4["A1"].alignment = CENTER
    ws4.row_dimensions[1].height = 28

    # Summary stats row
    ws4["A3"] = f"Critical: {len(seller_metrics[seller_metrics['risk_tier']=='CRITICAL'])}"
    ws4["C3"] = f"High: {len(seller_metrics[seller_metrics['risk_tier']=='HIGH'])}"
    ws4["E3"] = f"Predicted Anomalies (Next Month): {seller_metrics['predicted_anomaly_next_month'].sum()}"
    for col in ["A","C","E"]:
        ws4[f"{col}3"].font = Font(bold=True, size=10)

    risk_headers = ["Seller ID", "Seller Name", "Tier", "Category", "State",
                    "Disc. Rate %", "SLA Breaches", "Computed Risk Score",
                    "Risk Tier", "Predicted Anomaly?"]
    hdr_row = 5
    for ci, h in enumerate(risk_headers, 1):
        c = ws4.cell(row=hdr_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill(C_BLUE)
        c.alignment = CENTER
        c.border = thin_border()
        ws4.column_dimensions[get_column_letter(ci)].width = 20

    tier_colors_risk = {"CRITICAL": "D0021B", "HIGH": "F5A623", "MEDIUM": "FFFACD", "LOW": "D4EDDA"}
    tier_text_colors  = {"CRITICAL": C_WHITE,  "HIGH": C_WHITE,  "MEDIUM": "555500", "LOW": "005500"}

    display_sellers = seller_metrics.sort_values("computed_risk_score", ascending=False)

    for ri, (_, row) in enumerate(display_sellers.iterrows(), hdr_row + 1):
        rtier = row["risk_tier"]
        vals = [row["seller_id"], row["seller_name"][:30], row["tier"], row["category"],
                row["state"], round(row["disc_rate"], 2), row["sla_breaches"],
                round(row["computed_risk_score"], 1), rtier,
                "YES ⚠️" if row["predicted_anomaly_next_month"] else "No"]
        row_fill = tier_colors_risk[rtier] if rtier in ["CRITICAL","HIGH"] else ("F8F8F8" if ri % 2 == 0 else "FFFFFF")

        for ci, val in enumerate(vals, 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            if ci == 9:
                c.fill = header_fill(tier_colors_risk[rtier])
                c.font = Font(bold=True, color=tier_text_colors[rtier], size=10)
            elif ci == 10 and val == "YES ⚠️":
                c.fill = header_fill("FFF3CD")
                c.font = Font(bold=True, color="856404", size=10)
            else:
                c.fill = header_fill(row_fill)
                c.font = NORMAL_FONT
            c.alignment = CENTER
            c.border = thin_border()
        ws4.row_dimensions[ri].height = 17

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 5: CATEGORY PERFORMANCE
    # ─────────────────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Category Performance")
    ws5.sheet_properties.tabColor = C_GREEN

    ws5.merge_cells("A1:G2")
    ws5["A1"] = "📊  CATEGORY PERFORMANCE ANALYSIS"
    ws5["A1"].font = Font(bold=True, color=C_WHITE, size=14)
    ws5["A1"].fill = header_fill(C_DARK_GREEN)
    ws5["A1"].alignment = CENTER
    ws5.row_dimensions[1].height = 28

    cat_headers = ["Category", "Total GMV (₹)", "Transactions", "Net Payout (₹)",
                   "Discrepancies", "Disc. Rate %", "Avg Transaction (₹)"]
    for ci, h in enumerate(cat_headers, 1):
        c = ws5.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill(C_DARK_GREEN)
        c.alignment = CENTER
        c.border = thin_border()
        ws5.column_dimensions[get_column_letter(ci)].width = 22

    for ri, (_, row) in enumerate(category_data.iterrows(), 5):
        row_fill = "F0FFF0" if ri % 2 == 0 else "FFFFFF"
        vals = [row["category"], round(row["total_gmv"], 2), row["total_transactions"],
                round(row["total_payout"], 2), row["discrepancies"],
                round(row["disc_rate"], 2), round(row["avg_transaction"], 2)]
        for ci, val in enumerate(vals, 1):
            c = ws5.cell(row=ri, column=ci, value=val)
            c.fill = header_fill(row_fill)
            c.font = BOLD_FONT if ci == 1 else NORMAL_FONT
            c.alignment = LEFT if ci == 1 else CENTER
            c.border = thin_border()
            if ci in [2, 4, 7]:
                c.number_format = '₹#,##0.00'
        ws5.row_dimensions[ri].height = 18

        # Color disc rate
        disc_cell = ws5.cell(row=ri, column=6)
        dr = row["disc_rate"]
        if dr > 10:   disc_cell.font = Font(bold=True, color=C_RED, size=10)
        elif dr > 7:  disc_cell.font = Font(bold=True, color=C_AMBER, size=10)
        else:         disc_cell.font = Font(bold=True, color=C_GREEN, size=10)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 6: MANAGER ALERT DIGEST
    # ─────────────────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Manager Alert Digest")
    ws6.sheet_properties.tabColor = "9B59B6"

    ws6.merge_cells("A1:F2")
    ws6["A1"] = "⚡  MANAGER ALERT DIGEST — CRITICAL ITEMS ONLY"
    ws6["A1"].font = Font(bold=True, color=C_WHITE, size=14)
    ws6["A1"].fill = header_fill("6C3483")
    ws6["A1"].alignment = CENTER
    ws6.row_dimensions[1].height = 28

    critical_items = []

    # Critical root causes
    crit_rc = rc_summary[rc_summary["severity"] == "CRITICAL"]
    for _, r in crit_rc.iterrows():
        critical_items.append({
            "Priority": "P0 — CRITICAL",
            "Type": "Root Cause",
            "Description": f"{r['root_cause']}: {r['count']} cases totalling {inr(r['total_amount'])}",
            "Action Required": "Immediate investigation & fix",
            "Owner": "FinOps Lead",
            "Deadline": "24 hours",
        })

    # Sellers with critical risk
    crit_sellers = seller_metrics[seller_metrics["risk_tier"] == "CRITICAL"].head(5)
    for _, r in crit_sellers.iterrows():
        critical_items.append({
            "Priority": "P1 — HIGH",
            "Type": "Seller Risk",
            "Description": f"{r['seller_name'][:30]} (Risk Score: {r['computed_risk_score']}) — {r['discrepancy_count']} discrepancies",
            "Action Required": "Account review + payout hold",
            "Owner": "Seller Engagement",
            "Deadline": "48 hours",
        })

    # SLA aging
    old_breaches = aging_summary[aging_summary["aging_bucket"].str.contains("15+")]
    for _, r in old_breaches.iterrows():
        critical_items.append({
            "Priority": "P1 — HIGH",
            "Type": "SLA Breach",
            "Description": f"{r['count']} transactions outstanding 15+ days — {inr(r['amount_at_risk'])} at risk",
            "Action Required": "Escalate to resolution team",
            "Owner": "Operations",
            "Deadline": "Today",
        })

    alert_headers = ["Priority", "Type", "Description", "Action Required", "Owner", "Deadline"]
    priority_colors = {"P0 — CRITICAL": C_RED, "P1 — HIGH": "E67E22", "P2 — MEDIUM": "F1C40F"}

    for ci, h in enumerate(alert_headers, 1):
        c = ws6.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = header_fill("6C3483")
        c.alignment = CENTER
        c.border = thin_border()
        ws6.column_dimensions[get_column_letter(ci)].width = [14, 14, 55, 30, 18, 12][ci-1]

    for ri, item in enumerate(critical_items, 5):
        pri = item["Priority"]
        bg = priority_colors.get(pri, "FFFFFF")
        for ci, key in enumerate(alert_headers, 1):
            c = ws6.cell(row=ri, column=ci, value=item[key])
            if ci == 1:
                c.fill = header_fill(bg)
                c.font = Font(bold=True, color=C_WHITE if ci == 1 else C_AMAZON_DARK, size=9)
            else:
                c.fill = header_fill("FDF2FF" if ri % 2 == 0 else "FFFFFF")
                c.font = NORMAL_FONT
            c.alignment = WRAP if ci == 3 else CENTER
            c.border = thin_border()
        ws6.row_dimensions[ri].height = 30

    # ─────────────────────────────────────────────────────────────────────────
    # FINAL SAVE
    # ─────────────────────────────────────────────────────────────────────────
    fname = os.path.join(REPORT_DIR, "FinOps_Payout_Intelligence_Q4_2024.xlsx")
    wb.save(fname)
    print(f"   ✅ Excel report saved: {fname}")
    print(f"   📋 Sheets: Executive Summary · Root Cause Analysis · SLA Breach Tracker · Seller Risk Scores · Category Performance · Manager Alert Digest")
    return fname


# ─── GENERATE JSON FOR DASHBOARD ─────────────────────────────────────────────
def export_dashboard_data(monthly, rc_summary, aging_summary, seller_metrics, category_data, txn):
    print("\n📦 Exporting JSON data for dashboard...")

    disc = txn[txn["status"] == "Discrepancy"].copy()

    data = {
        "generated_at": datetime.now().isoformat(),
        "summary": {
            "total_gmv": round(txn["gmv"].sum(), 2),
            "total_payout": round(txn["actual_payout"].sum(), 2),
            "total_transactions": len(txn),
            "total_sellers": txn["seller_id"].nunique(),
            "discrepancy_count": int((txn["status"] == "Discrepancy").sum()),
            "amount_at_risk": round(disc["discrepancy_amount"].abs().sum(), 2),
            "sla_breaches": int(txn["sla_breach"].sum()),
            "discrepancy_rate": round((txn["status"] == "Discrepancy").mean() * 100, 2),
        },
        "monthly": monthly[["month","total_gmv","total_actual_payout","discrepancy_count",
                              "discrepancy_amount","sla_breaches","gmv_mom_growth"]].to_dict("records"),
        "root_causes": rc_summary.to_dict("records"),
        "aging": aging_summary.to_dict("records"),
        "categories": category_data[["category","total_gmv","total_transactions","disc_rate","discrepancies"]].to_dict("records"),
        "risk_tiers": seller_metrics.groupby("risk_tier")["seller_id"].count().to_dict(),
        "top_risky_sellers": seller_metrics.sort_values("computed_risk_score", ascending=False).head(10)[
            ["seller_id","seller_name","tier","computed_risk_score","risk_tier","disc_rate",
             "sla_breaches","predicted_anomaly_next_month","discrepancy_amount","total_gmv"]
        ].to_dict("records"),
        "recent_discrepancies": disc.sort_values("discrepancy_amount", key=abs, ascending=False).head(50)[
            ["transaction_id","seller_id","month","category","gmv","expected_payout",
             "actual_payout","discrepancy_amount","root_cause","days_outstanding",
             "resolution_status","sla_breach"]
        ].to_dict("records"),
        "seller_metrics": seller_metrics[[
            "seller_id","seller_name","tier","category","state","total_gmv",
            "disc_rate","computed_risk_score","risk_tier","sla_breaches",
            "predicted_anomaly_next_month","discrepancy_amount"
        ]].to_dict("records"),
    }

    with open("data/dashboard_data.json", "w") as f:
        json.dump(data, f, indent=2, default=str)
    print("   ✅ dashboard_data.json saved")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 65)
    print("  FINOPS SELLER PAYOUT INTELLIGENCE SYSTEM")
    print("  Amazon Seller Services · Q4 2024 Reconciliation")
    print("=" * 65)

    txn, sellers = load_data()
    print(f"\n📥 Loaded {len(txn):,} transactions · {len(sellers)} sellers from database")

    monthly          = monthly_reconciliation(txn, sellers)
    rc_summary, disc = root_cause_analysis(txn)
    aging_summary, critical_breaches, breaches = sla_tracker(txn, sellers)
    seller_metrics, high_risk = seller_risk_scoring(txn, sellers)
    category_data    = category_analysis(txn)

    build_excel_report(monthly, rc_summary, aging_summary, critical_breaches,
                       seller_metrics, high_risk, category_data, txn)
    export_dashboard_data(monthly, rc_summary, aging_summary, seller_metrics, category_data, txn)

    print("\n" + "=" * 65)
    print("  ✅ PIPELINE COMPLETE")
    print(f"  📊 Excel:     reports/FinOps_Payout_Intelligence_Q4_2024.xlsx")
    print(f"  🌐 Dashboard: open seller_payout_dashboard.html in browser")
    print("=" * 65)

if __name__ == "__main__":
    main()
